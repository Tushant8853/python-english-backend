#!/usr/bin/env python3
"""
Import lesson_library rows from an Excel workbook.

Usage (from python-backend/):
  python app/scripts/import_lesson_library_from_excel.py \\
    --file /path/to/English_Guru_100_Lesson_Production_Workbook.xlsx \\
    --dry-run

  python app/scripts/import_lesson_library_from_excel.py \\
    --file /path/to/English_Guru_100_Lesson_Production_Workbook.xlsx \\
    --upsert

Requires MONGODB_URI in .env or environment.
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.core.lesson_constants import LESSON_LEVELS, LESSON_TOPICS
from app.database.connection import close_database, connect_database
from app.database.indexes import ensure_lesson_library_indexes
from app.models.lesson_library import LessonDocument, QuizQuestion
from app.repositories.lesson_library_repository import LessonLibraryRepository

logger = logging.getLogger("english_guru.lesson_import")

REQUIRED_COLUMNS = ("Lesson ID", "Lesson Title", "Level")
PREFERRED_SHEETS = (
    "English Guru Roadmap",
    "Lessons",
    "Course Roadmap",
)
MCQ_COLUMNS = ("MCQ 1", "MCQ 2", "MCQ 3")
ALLOWED_LEVELS = frozenset(LESSON_LEVELS)
OPTION_LINE_RE = re.compile(r"^([A-Da-d])\)\s*(.+)$")
CORRECT_ANSWER_RE = re.compile(r"correct\s*answer\s*:\s*([A-Da-d])", re.IGNORECASE)
LESSON_CODE_NUMBER_RE = re.compile(r"(\d+)\s*$")
CHECKMARK_CHARS = ("✅", "✓", "☑", "(correct)", "[correct]")

TOPIC_KEYWORDS: tuple[tuple[str, str], ...] = (
    ("speaking", "conversation"),
    ("conversation", "conversation"),
    ("greeting", "conversation"),
    ("introduction", "conversation"),
    ("interview", "conversation"),
    ("grammar", "grammar"),
    ("vocabulary", "vocabulary"),
    ("listening", "listening"),
    ("reading", "reading"),
    ("writing", "grammar"),
)


@dataclass
class SkippedRow:
    row_number: int
    reason: str


@dataclass
class ImportSummary:
    total_rows: int = 0
    valid_lessons: int = 0
    would_create: int = 0
    would_update: int = 0
    created: int = 0
    updated: int = 0
    skipped: list[SkippedRow] = field(default_factory=list)
    malformed_mcqs: int = 0
    sheet_name: str = ""


@dataclass
class ParsedLesson:
    row_number: int
    lesson: LessonDocument


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )
    logging.getLogger("pymongo").setLevel(logging.WARNING)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_empty_row(values: tuple[Any, ...]) -> bool:
    return not any(cell_text(value) for value in values)


def select_worksheet(workbook: Workbook, sheet_name: str | None) -> Worksheet:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return workbook[sheet_name]

    for preferred in PREFERRED_SHEETS:
        if preferred in workbook.sheetnames:
            return workbook[preferred]

    return workbook[workbook.sheetnames[0]]


def read_header_map(worksheet: Worksheet) -> dict[str, int]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise ValueError("Worksheet is empty")

    headers = [cell_text(value) for value in first_row]
    header_map = {header: index for index, header in enumerate(headers) if header}
    missing = [column for column in REQUIRED_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return header_map


def row_value(row: tuple[Any, ...], header_map: dict[str, int], column: str) -> str:
    index = header_map.get(column)
    if index is None or index >= len(row):
        return ""
    return cell_text(row[index])


def day_number_from_lesson_code(lesson_code: str, row_order: int) -> int:
    match = LESSON_CODE_NUMBER_RE.search(lesson_code.strip().upper())
    if match:
        return int(match.group(1))
    return row_order


def topic_from_focus_area(focus_area: str) -> str:
    lowered = focus_area.strip().lower()
    if not lowered:
        return "conversation"

    for keyword, topic in TOPIC_KEYWORDS:
        if keyword in lowered:
            return topic

    return "conversation"


def build_description(row: tuple[Any, ...], header_map: dict[str, int]) -> str:
    sections: list[str] = []

    learning_objective = row_value(row, header_map, "Learning Objective")
    what_to_teach = row_value(row, header_map, "What To Teach")
    real_life = row_value(row, header_map, "Real-Life Scenario")
    practice_one = row_value(row, header_map, "Practice Activity 1")
    practice_two = row_value(row, header_map, "Practice Activity 2")
    assessment = row_value(row, header_map, "Assessment After Lesson?")
    notes = row_value(row, header_map, "Notes For Instructor")

    if learning_objective:
        sections.append(f"Learning Objective: {learning_objective}")
    if what_to_teach:
        sections.append(f"What To Teach: {what_to_teach}")
    if real_life:
        sections.append(f"Real-Life Scenario: {real_life}")

    practice_lines: list[str] = []
    if practice_one:
        practice_lines.append(f"1. {practice_one}")
    if practice_two:
        practice_lines.append(f"2. {practice_two}")
    if practice_lines:
        sections.append("Practice Activities:\n" + "\n".join(practice_lines))

    if assessment:
        sections.append(f"Assessment After Lesson: {assessment}")
    if notes:
        sections.append(f"Notes For Instructor: {notes}")

    return "\n\n".join(sections).strip()


def strip_checkmark(option_text: str) -> tuple[str, bool]:
    text = option_text.strip()
    marked = False
    for marker in CHECKMARK_CHARS:
        if marker in text:
            marked = True
            text = text.replace(marker, "").strip()
    return text, marked


def letter_to_index(letter: str) -> int | None:
    normalized = letter.strip().upper()
    if len(normalized) != 1 or normalized < "A" or normalized > "Z":
        return None
    return ord(normalized) - ord("A")


def parse_mcq_cell(raw_text: str) -> QuizQuestion | None:
    text = raw_text.strip()
    if not text:
        return None

    lines = [line.strip() for line in text.replace("\r\n", "\n").split("\n") if line.strip()]
    if not lines:
        return None

    options: list[str] = []
    marked_indices: list[int] = []
    question_lines: list[str] = []
    parsing_options = False

    for line in lines:
        option_match = OPTION_LINE_RE.match(line)
        if option_match:
            parsing_options = True
            option_body, marked = strip_checkmark(option_match.group(2))
            if not option_body:
                continue
            if marked:
                marked_indices.append(len(options))
            options.append(option_body)
            continue

        if not parsing_options:
            if line.lower().startswith("question:"):
                question_lines.append(line.split(":", 1)[1].strip())
            else:
                question_lines.append(line)

    question = " ".join(question_lines).strip()
    if question.lower().startswith("question:"):
        question = question.split(":", 1)[1].strip()

    if not question and options:
        question = "Choose the correct answer."

    correct_index: int | None = None
    if marked_indices:
        correct_index = marked_indices[0]
    else:
        joined = "\n".join(lines)
        answer_match = CORRECT_ANSWER_RE.search(joined)
        if answer_match:
            correct_index = letter_to_index(answer_match.group(1))

    if correct_index is None and options:
        return None
    if not question or len(options) < 2:
        return None
    if correct_index is None or correct_index < 0 or correct_index >= len(options):
        return None

    return QuizQuestion(question=question, options=options, correct_index=correct_index)


def parse_quizzes(
    row: tuple[Any, ...],
    header_map: dict[str, int],
    row_number: int,
    summary: ImportSummary,
) -> list[QuizQuestion]:
    questions: list[QuizQuestion] = []
    for column in MCQ_COLUMNS:
        if len(questions) >= 3:
            break
        raw = row_value(row, header_map, column)
        if not raw:
            continue
        parsed = parse_mcq_cell(raw)
        if parsed is None:
            summary.malformed_mcqs += 1
            logger.warning(
                "Malformed MCQ skipped",
                extra={"meta": {"row": row_number, "column": column}},
            )
            continue
        questions.append(parsed)
    return questions


def parse_workbook_rows(
    file_path: Path,
    sheet_name: str | None,
) -> tuple[list[ParsedLesson], ImportSummary]:
    if not file_path.is_file():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = select_worksheet(workbook, sheet_name)
        header_map = read_header_map(worksheet)
        summary = ImportSummary(sheet_name=worksheet.title)

        parsed_lessons: list[ParsedLesson] = []
        lesson_codes_seen: dict[str, int] = {}
        row_order = 0

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if is_empty_row(row):
                continue

            summary.total_rows += 1
            lesson_code_raw = row_value(row, header_map, "Lesson ID")
            title = row_value(row, header_map, "Lesson Title")
            level = row_value(row, header_map, "Level").upper()

            if not lesson_code_raw:
                summary.skipped.append(SkippedRow(row_number, "missing Lesson ID"))
                logger.warning("Skipped row %s: missing Lesson ID", row_number)
                continue

            if not title:
                summary.skipped.append(SkippedRow(row_number, "missing Lesson Title"))
                logger.warning("Skipped row %s (%s): missing Lesson Title", row_number, lesson_code_raw)
                continue

            lesson_code = lesson_code_raw.strip().upper()
            if lesson_code in lesson_codes_seen:
                raise ValueError(
                    f"Duplicate Lesson ID '{lesson_code}' in Excel "
                    f"(rows {lesson_codes_seen[lesson_code]} and {row_number})"
                )
            lesson_codes_seen[lesson_code] = row_number

            if level not in ALLOWED_LEVELS:
                summary.skipped.append(
                    SkippedRow(row_number, f"invalid Level '{level}' for {lesson_code}")
                )
                logger.warning(
                    "Skipped row %s (%s): invalid Level '%s'",
                    row_number,
                    lesson_code,
                    level,
                )
                continue

            row_order += 1
            focus_area = row_value(row, header_map, "Focus Area")
            lesson = LessonDocument(
                lesson_code=lesson_code,
                day_number=day_number_from_lesson_code(lesson_code, row_order),
                title=title,
                description=build_description(row, header_map),
                level=level,
                topic=topic_from_focus_area(focus_area),
                video_file_name="",
                external_video_url="",
                quiz=parse_quizzes(row, header_map, row_number, summary),
                is_active=True,
                sort_order=row_order,
            )
            parsed_lessons.append(ParsedLesson(row_number=row_number, lesson=lesson))
            summary.valid_lessons += 1
            logger.debug(
                "Parsed lesson %s (%s) with %s quiz items",
                lesson_code,
                row_number,
                len(lesson.quiz),
            )

        return parsed_lessons, summary
    finally:
        workbook.close()


async def apply_import(
    parsed_lessons: list[ParsedLesson],
    summary: ImportSummary,
    *,
    dry_run: bool,
) -> int:
    if dry_run:
        if not await connect_database():
            logger.info("Dry run: MongoDB unavailable; create/update counts assume no existing lessons")
            summary.would_create = summary.valid_lessons
            return 0

    if not await connect_database():
        logger.error("Could not connect to MongoDB. Set MONGODB_URI in .env")
        return 1

    await ensure_lesson_library_indexes()
    repo = LessonLibraryRepository()

    for item in parsed_lessons:
        lesson = item.lesson
        existing = await repo.find_by_lesson_code(lesson.lesson_code)

        if existing:
            summary.would_update += 1
            if dry_run:
                logger.debug("Would update %s (row %s)", lesson.lesson_code, item.row_number)
                continue

            day_conflict = await repo.find_by_day_number(lesson.day_number)
            if day_conflict and day_conflict.lesson_code != lesson.lesson_code:
                summary.skipped.append(
                    SkippedRow(
                        item.row_number,
                        f"dayNumber {lesson.day_number} already used by {day_conflict.lesson_code}",
                    )
                )
                logger.warning(
                    "Skipped update for %s: dayNumber %s already used by %s",
                    lesson.lesson_code,
                    lesson.day_number,
                    day_conflict.lesson_code,
                )
                continue

            existing.day_number = lesson.day_number
            existing.title = lesson.title
            existing.description = lesson.description
            existing.level = lesson.level
            existing.topic = lesson.topic
            existing.quiz = lesson.quiz
            existing.is_active = lesson.is_active
            existing.sort_order = lesson.sort_order
            await repo.save(existing)
            summary.updated += 1
            logger.debug("Updated lesson %s", lesson.lesson_code)
            continue

        summary.would_create += 1
        if dry_run:
            logger.debug("Would create %s (row %s)", lesson.lesson_code, item.row_number)
            continue

        day_conflict = await repo.find_by_day_number(lesson.day_number)
        if day_conflict:
            summary.skipped.append(
                SkippedRow(
                    item.row_number,
                    f"dayNumber {lesson.day_number} already used by {day_conflict.lesson_code}",
                )
            )
            logger.warning(
                "Skipped insert for %s: dayNumber %s already used by %s",
                lesson.lesson_code,
                lesson.day_number,
                day_conflict.lesson_code,
            )
            continue

        await repo.insert(lesson)
        summary.created += 1
        logger.debug("Inserted lesson %s", lesson.lesson_code)

    await close_database()
    return 0


def print_summary(summary: ImportSummary, *, dry_run: bool) -> None:
    mode = "DRY RUN" if dry_run else "IMPORT"
    logger.info("=== %s SUMMARY ===", mode)
    logger.info("Sheet: %s", summary.sheet_name)
    logger.info("Total non-empty rows: %s", summary.total_rows)
    logger.info("Valid lessons: %s", summary.valid_lessons)
    if dry_run:
        logger.info("Would create: %s", summary.would_create)
        logger.info("Would update: %s", summary.would_update)
    else:
        logger.info("Created: %s", summary.created)
        logger.info("Updated: %s", summary.updated)
    logger.info("Skipped rows: %s", len(summary.skipped))
    logger.info("Malformed MCQs skipped: %s", summary.malformed_mcqs)

    if summary.skipped:
        logger.info("Skipped row details:")
        for item in summary.skipped[:20]:
            logger.info("  row %s: %s", item.row_number, item.reason)
        if len(summary.skipped) > 20:
            logger.info("  ... and %s more", len(summary.skipped) - 20)


async def run(args: argparse.Namespace) -> int:
    file_path = Path(args.file).expanduser().resolve()
    dry_run = bool(args.dry_run)
    upsert = bool(args.upsert)

    if dry_run == upsert:
        logger.error("Specify exactly one of --dry-run or --upsert")
        return 1

    try:
        parsed_lessons, summary = parse_workbook_rows(file_path, args.sheet)
    except (FileNotFoundError, ValueError) as exc:
        logger.error("%s", exc)
        return 1

    exit_code = await apply_import(parsed_lessons, summary, dry_run=dry_run)
    print_summary(summary, dry_run=dry_run)

    if not dry_run and exit_code == 0:
        logger.info(
            "Video fields preserved on update (videoFileName, externalVideoUrl). "
            "New lessons imported with empty video fields."
        )

    return exit_code


def main() -> None:
    parser = argparse.ArgumentParser(description="Import lesson_library from Excel workbook")
    parser.add_argument(
        "--file",
        required=True,
        help="Path to Excel workbook (.xlsx)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name (default: preferred roadmap sheet or first sheet)",
    )
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true", help="Validate and print summary only")
    mode.add_argument("--upsert", action="store_true", help="Insert or update lessons in MongoDB")
    parser.add_argument("--verbose", action="store_true", help="Enable DEBUG logging")
    args = parser.parse_args()
    configure_logging(args.verbose)
    code = asyncio.run(run(args))
    sys.exit(code)


if __name__ == "__main__":
    main()
