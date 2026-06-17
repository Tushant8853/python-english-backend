#!/usr/bin/env python3
"""
Import placement_questions from Excel workbook.

Usage (from python-backend/):
  python app/scripts/import_placement_test_from_excel.py \\
    --file /path/to/English_Guru_30_Question_Placement_Test.xlsx \\
    --dry-run

  python app/scripts/import_placement_test_from_excel.py \\
    --file /path/to/English_Guru_30_Question_Placement_Test.xlsx \\
    --upsert

Requires MONGODB_URI in .env or environment.
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.database.connection import close_database, connect_database
from app.database.indexes import ensure_placement_question_indexes
from app.models.intake_question import IntakeOption
from app.models.placement_question import PlacementQuestionDocument, PlacementQuestionLocale
from app.repositories.placement_question_repository import PlacementQuestionRepository
from app.services.placement_onboarding_service import PlacementOnboardingService

logger = logging.getLogger("english_guru.placement_import")

REQUIRED_COLUMNS = (
    "Q No",
    "Question",
    "Option A",
    "Option B",
    "Option C",
    "Option D",
    "Correct Answer",
)
PREFERRED_SHEETS = ("30 Question Test", "Placement Test", "Questions")
OPTION_IDS = ("A", "B", "C", "D")
PROFILE_ANSWER = "profile"

# Learner-friendly Roman Hindi prompts and option labels keyed by Q No.
HINDI_CONTENT: dict[int, dict[str, Any]] = {
    1: {
        "prompt": "Apple kya hai:",
        "options": {"A": "Fal", "B": "Janwar", "C": "Gaadi", "D": "Jagah"},
    },
    2: {
        "prompt": "Doctor kahan kaam karta hai:",
        "options": {"A": "School", "B": "Hospital", "C": "Bank", "D": "Dukaan"},
    },
    3: {
        "prompt": "'Hot' ka opposite kya hai?",
        "options": {"A": "Garam", "B": "Thanda", "C": "Bada", "D": "Tez"},
    },
    4: {
        "prompt": "Kaun sa shabd action word hai?",
        "options": {"A": "Run (daudna)", "B": "Table", "C": "Blue", "D": "Happy"},
    },
    5: {
        "prompt": "Agar paise chahiye, aap kahan ja sakte ho:",
        "options": {"A": "Hospital", "B": "Bank", "C": "School", "D": "Kitchen"},
    },
    6: {
        "prompt": "I ___ a student.",
        "options": {"A": "am", "B": "is", "C": "are", "D": "be"},
    },
    7: {
        "prompt": "She ___ tea every morning.",
        "options": {"A": "drink", "B": "drinks", "C": "drinking", "D": "drank"},
    },
    8: {
        "prompt": "Yesterday, I ___ to the market.",
        "options": {"A": "go", "B": "goes", "C": "went", "D": "going"},
    },
    9: {
        "prompt": "They ___ playing outside now.",
        "options": {"A": "is", "B": "are", "C": "am", "D": "be"},
    },
    10: {
        "prompt": "The book is ___ the table.",
        "options": {"A": "on", "B": "in", "C": "at", "D": "to"},
    },
    11: {
        "prompt": "Padho: Rina subah 6 baje uthti hai. Woh chai banati hai aur bachche ko school ke liye taiyar karti hai. Rina kitne baje uthti hai?",
        "options": {"A": "5 AM", "B": "6 AM", "C": "7 AM", "D": "8 AM"},
    },
    12: {
        "prompt": "Padho: Amit apne maa-baap, wife aur do bachchon ke saath rehta hai. Woh ek chhoti dukaan mein kaam karta hai. Amit ke kitne bachche hain?",
        "options": {"A": "Ek", "B": "Do", "C": "Teen", "D": "Char"},
    },
    13: {
        "prompt": "Padho: Meena subah 8 baje nashta banati hai. Nashte ke baad woh ghar saaf karti hai aur 20 minute English padhti hai. Meena kya padhti hai?",
        "options": {"A": "Maths", "B": "English", "C": "Science", "D": "Hindi"},
    },
    14: {
        "prompt": "Padho: School meeting Monday 10 AM par hai. Parents ko apna ID card lana hoga. Parents ko kya lana chahiye?",
        "options": {"A": "Notebook", "B": "Lunch", "C": "ID card", "D": "Water bottle"},
    },
    15: {
        "prompt": "Padho: 'Please call me after 5 PM. I am busy in a meeting now.' Aap kab call karein?",
        "options": {"A": "Lunch se pehle", "B": "5 PM ke baad", "C": "10 AM par", "D": "Abhi"},
    },
    16: {
        "prompt": "Koi kehta hai: 'Good morning.' Sabse achha jawab kya hai?",
        "options": {"A": "Good morning", "B": "I am table", "C": "Yesterday", "D": "No food"},
    },
    17: {
        "prompt": "Apna introduction dene ke liye kaun sa sentence best hai?",
        "options": {
            "A": "My name is Rahul.",
            "B": "Rahul name my.",
            "C": "I Rahul name.",
            "D": "Name Rahul is my.",
        },
    },
    18: {
        "prompt": "Aapko madad chahiye. Aap kya kahenge?",
        "options": {
            "A": "Help me now!",
            "B": "Can you please help me?",
            "C": "You help.",
            "D": "Why help?",
        },
    },
    19: {
        "prompt": "Aap chahte ho bachcha padhai kare. Aap kya keh sakte ho?",
        "options": {
            "A": "Please study now.",
            "B": "Study you now please.",
            "C": "Now studying you.",
            "D": "You are study.",
        },
    },
    20: {
        "prompt": "Interview mein interviewer kehta hai: 'Tell me about yourself.' Shuruat ke liye best kya hai?",
        "options": {
            "A": "My name is Ravi and I am a graduate.",
            "B": "Food is good.",
            "C": "I am going market.",
            "D": "Yesterday rain.",
        },
    },
    21: {
        "prompt": "Sahi sentence chuniye:",
        "options": {
            "A": "I English study every day.",
            "B": "I study English every day.",
            "C": "Every day English I study.",
            "D": "Study I English every day.",
        },
    },
    22: {
        "prompt": "Sahi sentence chuniye:",
        "options": {
            "A": "She goes to school.",
            "B": "She go to school.",
            "C": "She going school.",
            "D": "She school goes to.",
        },
    },
    23: {
        "prompt": "Sahi question chuniye:",
        "options": {
            "A": "Where you live?",
            "B": "Where do you live?",
            "C": "Where live you?",
            "D": "Where are live?",
        },
    },
    24: {
        "prompt": "Sahi sentence chuniye:",
        "options": {
            "A": "I went to work yesterday.",
            "B": "I go to work yesterday.",
            "C": "I going work yesterday.",
            "D": "I was go work yesterday.",
        },
    },
    25: {
        "prompt": "Sahi sentence chuniye:",
        "options": {
            "A": "I will call you tomorrow.",
            "B": "I call you yesterday tomorrow.",
            "C": "I called you tomorrow.",
            "D": "I calling you tomorrow.",
        },
    },
    26: {
        "prompt": "Aap English mainly kyun seekhna chahte ho?",
        "options": {
            "A": "Bachche / family se baat karne ke liye",
            "B": "Job ya interview ke liye",
            "C": "Roz confidence ke liye",
            "D": "Padhai / school ke liye",
        },
    },
    27: {
        "prompt": "English bolte waqt aap kitne confident feel karte ho?",
        "options": {"A": "Bahut kam", "B": "Kam", "C": "Medium", "D": "Zyada"},
    },
    28: {
        "prompt": "Kya aap bolne se pehle dimaag mein Hindi se English translate karte ho?",
        "options": {"A": "Hamesha", "B": "Kabhi-kabhi", "C": "Kam hi", "D": "Kabhi nahi"},
    },
    29: {
        "prompt": "Roz kitna time padhai kar sakte ho?",
        "options": {"A": "10 minute", "B": "20 minute", "C": "30 minute", "D": "45+ minute"},
    },
    30: {
        "prompt": "Aapko sabse zyada kaun describe karta hai?",
        "options": {
            "A": "Parent / Housewife",
            "B": "Job seeker",
            "C": "Student",
            "D": "Working professional",
        },
    },
}


@dataclass
class RowError:
    row_number: int
    reason: str


@dataclass
class ImportSummary:
    sheet_name: str = ""
    total_rows: int = 0
    valid_questions: int = 0
    profile_questions: int = 0
    invalid_rows: int = 0
    would_create: int = 0
    would_update: int = 0
    created: int = 0
    updated: int = 0
    would_remove_legacy: int = 0
    removed_legacy: int = 0
    errors: list[RowError] = field(default_factory=list)


@dataclass
class ParsedQuestion:
    row_number: int
    question: PlacementQuestionDocument


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )
    logging.getLogger("pymongo").setLevel(logging.WARNING)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_empty_row(values: tuple[Any, ...]) -> bool:
    return not any(cell_text(value) for value in values)


def select_worksheet(workbook: Workbook, sheet_name: str | None) -> Worksheet:
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return workbook[sheet_name]

    for preferred in PREFERRED_SHEETS:
        if preferred in workbook.sheetnames:
            return workbook[preferred]

    for name in workbook.sheetnames:
        ws = workbook[name]
        headers = [cell_text(value) for value in next(ws.iter_rows(max_row=1, values_only=True), ())]
        if all(column in headers for column in REQUIRED_COLUMNS):
            return ws

    return workbook[workbook.sheetnames[0]]


def read_header_map(worksheet: Worksheet) -> dict[str, int]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise ValueError("Worksheet is empty")

    headers = [cell_text(value) for value in first_row]
    header_map = {header: index for index, header in enumerate(headers) if header}
    missing = [column for column in REQUIRED_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return header_map


def row_value(row: tuple[Any, ...], header_map: dict[str, int], column: str) -> str:
    index = header_map.get(column)
    if index is None or index >= len(row):
        return ""
    return cell_text(row[index])


def question_key_for_order(order: int) -> str:
    return f"PLACEMENT_Q_{order:03d}"


def normalize_correct_answer(raw: str) -> str | None:
    value = raw.strip().upper()
    if not value:
        return None
    if value == PROFILE_ANSWER.upper():
        return PROFILE_ANSWER
    if value in OPTION_IDS:
        return value
    return None


def build_locale(prompt: str, options: dict[str, str], language: str, order: int) -> PlacementQuestionLocale:
    if language == "hi":
        hi = HINDI_CONTENT.get(order, {})
        hi_prompt = str(hi.get("prompt") or prompt)
        hi_options = hi.get("options") or {}
        option_labels = [str(hi_options.get(option_id, options[option_id])) for option_id in OPTION_IDS]
        prompt_text = hi_prompt
    else:
        option_labels = [options[option_id] for option_id in OPTION_IDS]
        prompt_text = prompt

    return PlacementQuestionLocale(
        prompt=prompt_text,
        options=[
            IntakeOption(id=option_id, label=option_labels[index])
            for index, option_id in enumerate(OPTION_IDS)
        ],
    )


def parse_workbook(file_path: Path, sheet_name: str | None) -> tuple[list[ParsedQuestion], ImportSummary]:
    if not file_path.is_file():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = select_worksheet(workbook, sheet_name)
        header_map = read_header_map(worksheet)
        summary = ImportSummary(sheet_name=worksheet.title)
        parsed: list[ParsedQuestion] = []
        seen_orders: dict[int, int] = {}

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if is_empty_row(row):
                continue

            summary.total_rows += 1
            q_no_raw = row_value(row, header_map, "Q No")
            if not q_no_raw:
                summary.invalid_rows += 1
                summary.errors.append(RowError(row_number, "missing Q No"))
                continue

            try:
                order = int(float(q_no_raw))
            except ValueError:
                summary.invalid_rows += 1
                summary.errors.append(RowError(row_number, f"invalid Q No '{q_no_raw}'"))
                continue

            if order in seen_orders:
                raise ValueError(
                    f"Duplicate Q No {order} in Excel (rows {seen_orders[order]} and {row_number})"
                )
            seen_orders[order] = row_number

            prompt = row_value(row, header_map, "Question")
            if not prompt:
                summary.invalid_rows += 1
                summary.errors.append(RowError(row_number, f"Q{order}: empty Question"))
                continue

            options = {
                option_id: row_value(row, header_map, f"Option {option_id}")
                for option_id in OPTION_IDS
            }
            missing_options = [option_id for option_id, label in options.items() if not label]
            if missing_options:
                summary.invalid_rows += 1
                summary.errors.append(
                    RowError(row_number, f"Q{order}: missing options {', '.join(missing_options)}")
                )
                continue

            correct_raw = row_value(row, header_map, "Correct Answer")
            correct_answer = normalize_correct_answer(correct_raw)
            if correct_answer is None:
                summary.invalid_rows += 1
                summary.errors.append(
                    RowError(row_number, f"Q{order}: invalid Correct Answer '{correct_raw}'")
                )
                continue

            is_profile = correct_answer == PROFILE_ANSWER
            if is_profile:
                summary.profile_questions += 1
                correct_option_id = None
            else:
                correct_option_id = correct_answer

            content = {
                "en": build_locale(prompt, options, "en", order),
                "hi": build_locale(prompt, options, "hi", order),
            }
            question = PlacementQuestionDocument(
                question_key=question_key_for_order(order),
                order=order,
                correct_option_id=correct_option_id,
                is_active=True,
                content=content,
            )
            parsed.append(ParsedQuestion(row_number=row_number, question=question))
            summary.valid_questions += 1
            logger.debug(
                "Parsed %s order=%s profile=%s",
                question.question_key,
                order,
                is_profile,
            )

        if summary.invalid_rows > 0:
            raise ValueError(
                f"Validation failed: {summary.invalid_rows} invalid row(s). Fix Excel and retry."
            )

        return parsed, summary
    finally:
        workbook.close()


async def count_legacy_to_remove(
    repo: PlacementQuestionRepository,
    imported_keys: set[str],
) -> int:
    all_questions = await repo.list_all(include_inactive=True)
    return sum(1 for question in all_questions if question.question_key not in imported_keys)


async def remove_legacy_questions(
    repo: PlacementQuestionRepository,
    imported_keys: set[str],
) -> int:
    all_questions = await repo.list_all(include_inactive=True)
    removed = 0
    for question in all_questions:
        if question.question_key in imported_keys:
            continue
        if question.id and await repo.delete(question.id):
            removed += 1
            logger.debug("Removed legacy question %s", question.question_key)
    return removed


async def apply_import(
    parsed_questions: list[ParsedQuestion],
    summary: ImportSummary,
    *,
    dry_run: bool,
) -> int:
    imported_keys = {item.question.question_key for item in parsed_questions}

    if dry_run:
        if await connect_database():
            repo = PlacementQuestionRepository()
            for item in parsed_questions:
                existing = await repo.find_by_key(item.question.question_key)
                if existing:
                    summary.would_update += 1
                else:
                    summary.would_create += 1
            summary.would_remove_legacy = await count_legacy_to_remove(repo, imported_keys)
            await close_database()
        else:
            summary.would_create = summary.valid_questions
        return 0

    if not await connect_database():
        logger.error("Could not connect to MongoDB. Set MONGODB_URI in .env")
        return 1

    await ensure_placement_question_indexes()
    repo = PlacementQuestionRepository()

    for item in parsed_questions:
        question = item.question
        existing = await repo.find_by_key(question.question_key)
        if existing:
            question.id = existing.id
            question.created_at = existing.created_at
            await repo.replace(question)
            summary.updated += 1
            logger.debug("Updated %s", question.question_key)
        else:
            await repo.create(question)
            summary.created += 1
            logger.debug("Created %s", question.question_key)

    summary.removed_legacy = await remove_legacy_questions(repo, imported_keys)
    await close_database()
    return 0


async def verify_placement_endpoint() -> None:
    if not await connect_database():
        logger.warning("Skipping endpoint verification: MongoDB unavailable")
        return

    service = PlacementOnboardingService()
    payload = await service.get_questions("en")
    questions = payload.get("questions") or []
    logger.info("Verification: %s active questions returned (en)", len(questions))
    if questions:
        orders = [int(q.get("order") or 0) for q in questions]
        logger.info("Verification: order range %s–%s", min(orders), max(orders))
        logger.info(
            "Verification: first=%s last=%s",
            questions[0].get("questionKey"),
            questions[-1].get("questionKey"),
        )
        profile_count = sum(
            1
            for item in questions
            if str(item.get("questionKey", "")).startswith("PLACEMENT_Q_")
            and int(item.get("order") or 0) >= 26
        )
        logger.info("Verification: profile questions in response (order 26+): %s", profile_count)

    hi_payload = await service.get_questions("hi")
    hi_questions = hi_payload.get("questions") or []
    if hi_questions:
        sample = hi_questions[2]
        logger.info(
            "Verification: hi sample Q3 prompt=%s",
            str(sample.get("prompt", ""))[:60],
        )

    await close_database()


def print_summary(summary: ImportSummary, *, dry_run: bool) -> None:
    mode = "DRY RUN" if dry_run else "IMPORT"
    logger.info("=== %s SUMMARY ===", mode)
    logger.info("Sheet: %s", summary.sheet_name)
    logger.info("Total rows: %s", summary.total_rows)
    logger.info("Valid questions: %s", summary.valid_questions)
    logger.info("Profile questions: %s", summary.profile_questions)
    logger.info("Invalid rows: %s", summary.invalid_rows)
    if dry_run:
        logger.info("Would create: %s", summary.would_create)
        logger.info("Would update: %s", summary.would_update)
        logger.info("Would remove legacy questions: %s", summary.would_remove_legacy)
    else:
        logger.info("Created: %s", summary.created)
        logger.info("Updated: %s", summary.updated)
        logger.info("Removed legacy questions: %s", summary.removed_legacy)

    if summary.errors:
        logger.info("Invalid row details:")
        for item in summary.errors[:20]:
            logger.info("  row %s: %s", item.row_number, item.reason)


async def run(args: argparse.Namespace) -> int:
    file_path = Path(args.file).expanduser().resolve()
    dry_run = bool(args.dry_run)
    upsert = bool(args.upsert)

    if dry_run == upsert:
        logger.error("Specify exactly one of --dry-run or --upsert")
        return 1

    try:
        parsed_questions, summary = parse_workbook(file_path, args.sheet)
    except (FileNotFoundError, ValueError) as exc:
        logger.error("%s", exc)
        return 1

    exit_code = await apply_import(parsed_questions, summary, dry_run=dry_run)
    print_summary(summary, dry_run=dry_run)

    if not dry_run and exit_code == 0:
        await verify_placement_endpoint()

    return exit_code


def main() -> None:
    parser = argparse.ArgumentParser(description="Import placement_questions from Excel workbook")
    parser.add_argument("--file", required=True, help="Path to Excel workbook (.xlsx)")
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: auto-detect)")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true", help="Validate and print summary only")
    mode.add_argument("--upsert", action="store_true", help="Insert or update questions in MongoDB")
    parser.add_argument("--verbose", action="store_true", help="Enable DEBUG logging")
    args = parser.parse_args()
    configure_logging(args.verbose)
    code = asyncio.run(run(args))
    sys.exit(code)


if __name__ == "__main__":
    main()
